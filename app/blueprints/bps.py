from app.models.tables import AnexosFuncionario    
from functools import wraps
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
import os
from flask import Blueprint, redirect, render_template, request, session, url_for, send_file, abort
from app.auth.auth import auth, get_google_provider_cfg, client, GOOGLE_CLIENT_SECRET, GOOGLE_CLIENT_ID
from app.controllers.TipoProcessoController import TipoProcessoController
from app.controllers.UserController import UserController
from app.controllers.DocsControlller import DocsController
from app.controllers.EmpresaController import EmpresaController
from app.controllers.CuboController import CuboController
from app.controllers.CategoriaController import CategoriaController
from app.controllers.ContratoController import ContratoController
from app.controllers.PerfilController import PerfilController
from app.controllers.LogController import LogController
from app.controllers.IntegraController import IntegraController

from app.ext.db import db

from io import BytesIO
import urllib
import random
import requests
import json
import zlib

from app.models.tables import Anexo, Aprovacao, Contrato, Empresa, StatusFuncionario, Subcont

from functools import wraps
from datetime import datetime, timedelta
from flask import abort, redirect, url_for, session, request, current_app
from werkzeug.utils import secure_filename
import secrets
from flask import jsonify

bp_app = Blueprint("bp", __name__)

def init_session_security(app):
    """Configure secure session settings"""
    app.config.update(
        SESSION_COOKIE_SECURE=True,
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SAMESITE="Lax",
        PERMANENT_SESSION_LIFETIME=timedelta(hours=24),
        SESSION_PROTECTION="strong"
    )

def validate_session():
    """Validate session security and freshness"""
    if "last_activity" in session:
        inactivity_period = datetime.utcnow() - datetime.fromisoformat(session["last_activity"])
        if inactivity_period > timedelta(hours=24):
            session.clear()
            return False
    session["last_activity"] = datetime.utcnow().isoformat()
    if "session_token" not in session:
        session["session_token"] = secrets.token_urlsafe(32)
    return True

def requires_prestadora_auth(f):
    """Decorator for prestadora authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "chave_empresa" not in session:
            return redirect(url_for("bp.login_prestadora"))
        if not validate_session():
            return redirect(url_for("bp.login_prestadora"))
        return f(*args, **kwargs)
    return decorated_function

def permission_required(required_permission=None):
    """Enhanced permission decorator with session validation"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if "id" not in session or not validate_session():
                session.clear()
                return redirect(url_for("bp.index"))
            try:
                user = UserController.get(session["id"])
                if not user:
                    session.clear()
                    return redirect(url_for("bp.index"))
                if not user.perfil:
                    return redirect(url_for("bp.blocked"))
                if "user_permissions" not in session:
                    permissao = PerfilController.get(user.perfil_id)
                    session["user_permissions"] = {
                        attr: getattr(permissao, attr)
                        for attr in dir(permissao)
                        if not attr.startswith("_") and attr not in ["metadata", "query", "registry"]
                    }
                    
                if required_permission:
                    if not session["user_permissions"].get(required_permission, False):
                        abort(403)
                return f(*args, **kwargs)
            except Exception as e:
                current_app.logger.error(f"Permission check error: {str(e)}")
                session.clear()
                return redirect(url_for("bp.index"))
        return decorated_function
    return decorator

@bp_app.route("/")
def index():
    return render_template("login.html")

@bp_app.route("/login")
def login():
    if "id" in session.keys():
        return redirect(url_for("bp.home"))
    else:
        request_uri = auth()
        return redirect(request_uri)

@bp_app.route("/login/get_user_info")
def get_user_info():
    code = request.args.get("code")
    google_provider_cfg = get_google_provider_cfg()
    token_endpoint = google_provider_cfg["token_endpoint"]
    token_url, headers, body = client.prepare_token_request(
        token_endpoint,
        authorization_response=request.url,
        redirect_url=request.base_url,
        code=code,
    )
    
    token_response = requests.post(
        token_url,
        headers=headers,
        data=body,
        auth=(GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET),
    )
    client.parse_request_body_response(json.dumps(token_response.json()))
    client.access_token
    session["google_tokens"] = {
        "access_token": client.access_token,
        "refresh_token": client.refresh_token,
        "expires_in": client.expires_in
    }
    
    userinfo_endpoint = google_provider_cfg["userinfo_endpoint"]
    uri, headers, body = client.add_token(userinfo_endpoint)
    userinfo_response = requests.get(uri, headers=headers, data=body)
    if userinfo_response.json().get("email_verified"):
        unique_id = userinfo_response.json()["sub"]
        users_email = userinfo_response.json()["email"]
        picture = userinfo_response.json()["picture"]
        users_name = userinfo_response.json()["given_name"]
    else:
        return "User email not available or not verified by Google.", 400
    user = {
        "id": unique_id, "nome": users_name, "email": users_email, "icon": picture
    }
    icon = user["icon"].split("/")[-1]
    return redirect(url_for("bp.create_user", _id=user["id"], email=user["email"], nome=user["nome"], icon=icon))

@bp_app.route("/create_user/<_id>/<email>/<nome>/<icon>", methods=["GET"])
def create_user(_id, email, nome, icon):
    try:
        user = UserController.get(_id)
        if user:
            session.permanent = True
            session["id"] = user._id
            session["nome"] = user.nome
            session["email"] = user.email
            session["perfil"] = user.perfil
            session["last_activity"] = datetime.utcnow().isoformat()
            session["session_token"] = secrets.token_urlsafe(32)
            LogController.create(session["nome"],
                              session["perfil"],
                              "AUTH",
                              "LOGIN",
                              "User logged in successfully")
            if user.perfil:
                permissao = PerfilController.get(user.perfil_id)
                permissions = {}
                for attr in dir(permissao):
                    if not attr.startswith("_") and attr not in ["metadata", "query", "registry"]:
                        value = getattr(permissao, attr)
                        if isinstance(value, (bool, int, float, str)) or value is None:
                            permissions[attr] = value
                session["user_permissions"] = permissions
            return redirect("/home")
        else:
            user = UserController.create(
                _id=_id,
                file=f"https://lh3.googleusercontent.com/a/{icon}",
                nome=nome,
                email=email
            )
            
            session.permanent = True
            session["id"] = user._id
            session["nome"] = user.nome
            session["email"] = user.email
            session["perfil"] = user.perfil
            session["last_activity"] = datetime.utcnow().isoformat()
            session["session_token"] = secrets.token_urlsafe(32)
            LogController.create(session["nome"],
                              session["perfil"],
                              "USER",
                              "CREATE",
                              "New user account created")
            return redirect(url_for("bp.home"))
        
    except Exception as e:
        current_app.logger.error(f"User creation/login error: {str(e)}")
        session.clear()
        return redirect(url_for("bp.index"))

@bp_app.route("/logout")
def logout():
    """Secure logout handling"""
    try:
        user_id = session.get("id")
        if user_id:
            LogController.create(session.get("nome", "Unknown"),
                              session.get("perfil", "Unknown"),
                              "AUTH",
                              "LOGOUT",
                              "User logged out successfully")
        session.clear()
        return redirect(url_for("bp.index"))
    except Exception as e:
        current_app.logger.error(f"Logout error: {str(e)}")
        session.clear()
        return redirect(url_for("bp.index"))

@bp_app.route("/logout_prestadora")
def logout_prestadora():
    """Secure logout handling for prestadora"""
    try:
        empresa_nome = session.get("empresa_nome", "Unknown")
        if empresa_nome:
            LogController.create(empresa_nome,
                              session.get("perfil", "Unknown"),
                              "AUTH_PRESTADORA",
                              "LOGOUT",
                              "Prestadora logged out successfully")
    
        session.clear()
        return redirect(url_for("bp.login_prestadora"))
    except Exception as e:
        current_app.logger.error(f"Prestadora logout error: {str(e)}")
        session.clear()
        return redirect(url_for("bp.login_prestadora"))

@bp_app.route("/blocked")
def blocked():
    return render_template("blocked.html")

@bp_app.route("/home")
def home():
    if "id" in session.keys():
        user = UserController.get(session["id"])
        if not user.perfil:
            return redirect(url_for("bp.blocked"))
        perfil_id = UserController.get(session["id"]).perfil_id
        permissao = PerfilController.get(perfil_id)
        return render_template("index.html", user=user, permissao=permissao)
    else:
        return redirect(url_for("bp.index"))

@bp_app.route("/index_prestadora")
def index_prestadora():
    chave = session["chave_empresa"]
    prestadora = EmpresaController.get(chave=chave)
    return render_template("index_prestadora.html", prestadora=prestadora)

@bp_app.route("/upload_file_view")
def upload_file_view():
    chave = session["chave_empresa"]
    empresa_id = EmpresaController.get(chave=chave)._id
    contratos = Contrato.query.filter_by(empresa_id=empresa_id).all()
    categorias = CategoriaController.get_all()
    tipo_processo = TipoProcessoController.get_all()
    return render_template("upload_file.html", contratos=contratos, categorias=categorias, tipo_processo=tipo_processo)

@bp_app.route("/upload_file_action", methods=["POST"])
def upload_file_action():
    form = request.form
    files = request.files.getlist("anexo")
    ALLOWED_EXTENSIONS = {".pdf", ".xlsx"}
    
    print(form)
    
    for file in files:
        if file and file.filename:
            file_ext = os.path.splitext(secure_filename(file.filename))[1].lower()
            if file_ext not in ALLOWED_EXTENSIONS:
                return "Error: apenas .pdf e .xlsx são permitidos", 400
            
    chave = session["chave_empresa"]
    empresa_id = EmpresaController.get(chave=chave)._id
    categoria_nome = CategoriaController.get(form["categoria"].split("|")[0]).nome
    contrato_nome = ContratoController.get(form["contrato"]).nome
    form = dict(form)
    
    form["empresa"] = empresa_id
    form["categoria_nome"] = categoria_nome
    form["contrato_nome"] = contrato_nome
    DocsController.create(form, files)
    
    return "OK"

@bp_app.route("/status_files/")
def status_files():
    chave = session["chave_empresa"]
    empresa_id = EmpresaController.get(chave=chave)._id
    documentos_empresa = list(DocsController.get_all(emp=empresa_id).items())
    return render_template("status_documentos.html", docs_emp=documentos_empresa)

@bp_app.route("/auth_prestadora", methods=["POST"])
def auth_prestadora():
    try:
        form = request.form
        chave = form["chave"]
        nome = form["nome"].upper()
        empresa = EmpresaController.auth(nome, chave)
        
        if not empresa:
            print(empresa)
            return "Erro: Empresa não encontrada ou chave inválida"
        
        session.permanent = True
        session["chave_empresa"] = chave
        session["empresa_nome"] = nome
        session["nome"] = nome
        session["perfil"] = f"PRESTADORA: {nome}"
        session["last_activity"] = datetime.utcnow().isoformat()
        session["session_token"] = secrets.token_urlsafe(32)
        
        LogController.create(session["empresa_nome"],
                           session["perfil"],
                           "AUTH_PRESTADORA",
                           "LOGIN",
                           f"Prestadora {nome} logged in successfully")
        print(f"Prestadora {nome} authenticated successfully")
        return "ok"
    
    except Exception as e:
        print(f"Error during prestadora authentication: {str(e)}")
        
        current_app.logger.error(f"Prestadora authentication error: {str(e)}")
        session.clear()
        return f"erro: {str(e)}"

@bp_app.route("/anexo/<id>")
def serve_anexo(id):
    anexo = Anexo.query.filter_by(_id=id).first()
    if anexo is None:
        abort(404, description="Documento não encontrado")
    try:
        decompressed_data = zlib.decompress(anexo.data)
    except zlib.error:
        abort(500, description="Erro ao descompactar o documento")
    temp_file = BytesIO(decompressed_data)
    
    return send_file(temp_file, download_name=f"{anexo.filename}.pdf")

@bp_app.route("/login_prestadora")
def login_prestadora():
    return render_template("login_prestadora.html")

@bp_app.route("/gen_key")
def generate_key():
    return str(random.randint(100000, 999999))

@bp_app.route("/create_empresa/<nome>/<chave>/<cnpj>/<dep>/<status>/")
@bp_app.route("/create_empresa/<nome>/<chave>/<cnpj>/<dep>/<status>/<_id>")
def create_empresa(nome, chave, cnpj, dep, status, _id=None):
    cnpj = cnpj.replace("&", "/")
    
    try:
        if _id:
            EmpresaController.update(nome, chave, cnpj, dep, status, _id)
        else:
            EmpresaController.create(nome, chave, cnpj, dep, status)
        return "ok"
    except Exception as e:
        return f"erro {e}"

@bp_app.route("/exclui_empresa/<_id>")
def exclui_empresa(_id):
    try:
        EmpresaController.delete(_id=_id)
        return "ok"
    except Exception as e:
        return f"erro {e}"

@bp_app.route("/list_empresas/")
def list_empresas():
    empresas = EmpresaController.get_all()
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("list_empresas.html", empresas=empresas, permissao=permissao)

@bp_app.route("/filter_prestadoras/")
def list_prestadoras_filter():
    docs = DocsController.get_all()
    print(f"Total de documentos: {len(docs)}")
    
    docs = docs.items()
    return render_template("list_prestadoras_filter.html", docs=docs)

@bp_app.route("/filter_docs/<int:empresa_id>/", defaults={"content": None})
@bp_app.route("/filter_docs/<int:empresa_id>/<path:content>")
def list_docs_history(empresa_id, content):
    content = urllib.parse.unquote(content) if content else ""
    hist = False
    
    if "hist" in content:
        titulo = content.split(",")[1]
        docs = DocsController.get_history_docs(titulo)
        hist = True
    else:
        docs = DocsController.filter(empresa_id, content)
    docs = docs.items()
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    
    print(docs)
    
    return render_template("list_docs_filter.html", docs_emp=docs, permissao=permissao, hist=hist)

@bp_app.route("/filter_docs_prestadora/<int:empresa_id>/", defaults={"content": None})
@bp_app.route("/filter_docs_prestadora/<int:empresa_id>/<path:content>")
def list_docs_filter_prestadora(empresa_id, content):
    content = urllib.parse.unquote(content) if content else ""
    hist = False
    
    if "hist" in content:
        titulo = content.split(",")[1]
        docs = DocsController.get_history_docs(titulo)
        hist = True
    else:
        docs = DocsController.filter(empresa_id, content)
    docs = docs.items()
    return render_template("status_documentos.html", docs_emp=docs, hist=hist)

@bp_app.route("/list_docs/<int:empresa_id>")
def list_docs_filter(empresa_id):
    hist = False
    docs = DocsController.get_all(empresa_id)
    docs = docs.items()
    
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("list_docs_filter.html", docs_emp=docs, permissao=permissao, hist=hist)

@bp_app.route("/list_docs_prestadora/<int:empresa_id>")
def list_docs_prestadora(empresa_id):
    hist = False
    docs = DocsController.get_all(empresa_id)
    docs = docs.items()
    return render_template("status_documentos.html", docs_emp=docs, hist=hist)

@bp_app.route("/update_documento/", methods=["POST"])
def update_documento():
    files = request.files
    form = request.form
    DocsController.corrige_documento(form["id"], files, form=form)
    return "ok"

@bp_app.route("/exclui_documentos/")
def delete_documento():
    try:
        ids = request.args.get("ids", "")
        ids = ids.split(",") if ids else []
        for id in ids:
            DocsController.delete(id)
        
        return "ok"
    except Exception as e:
        return f"erro {e}"

@bp_app.route("/cubo_menu")
def cubo_menu():
    tipos_processos = TipoProcessoController.get_all()
    contratos = ContratoController.get_all()
    perfis = PerfilController.get_all()
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("menu_cubo.html", tipos_processos=tipos_processos,
                           contratos=contratos,
                           perfis=perfis,
                           permissao=permissao)

@bp_app.route("/create_cubo", methods=["POST"])
def create_cubo():
    form = dict(request.form)
    form["perfil_nome"] = PerfilController.get(form["perfil"]).nome
    try:
        if form["_id"]:
            CuboController.update(form)
        else:
            CuboController.create(form)
        return "ok"
    except Exception as e:
        return (f"erro:{e}")

@bp_app.route("/delete_cubo/<_id>")
def delete_cubo(_id):
    try:
        CuboController.delete(_id)
        return "ok"
    except Exception as e:
        return (f"erro:{e}")

@bp_app.route("/list_cubo/")
def list_cubo():
    cubos = CuboController.get_all()
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("list_cubo.html", cubos=cubos, permissao=permissao)

@bp_app.route("/contrato_menu")
def contrato_menu():
    empresas = EmpresaController.get_all()
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("menu_contrato.html", empresas=empresas, permissao=permissao)

@bp_app.route("/create_contrato", methods=["POST"])
def create_contrato():
        form = dict(request.form)
        
        if form["_id"]:
            ContratoController.update(form)
            print("Contrato atualizado com sucesso")
        else:
            form = dict(form)
            form["empresa_nome"] = EmpresaController.get(_id=form["empresa"]).nome
            ContratoController.create(form)
            
            print("Contrato criado com sucesso")
        return "ok"

@bp_app.route("/delete_contrato/<_id>")
def delete_contrato(_id):
    try:
        ContratoController.delete(_id)
        return "ok"
    except Exception as e:
        return (f"erro:{e}")

@bp_app.route("/list_contratos/")
def list_contratos():
    contratos = ContratoController.get_all()
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("list_contratos.html", contratos=contratos, permissao=permissao)

@bp_app.route("/categorias_menu")
def categorias_menu():
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    tipos_processos = TipoProcessoController.get_all()
    return render_template("menu_categorias.html", permissao=permissao, tipos_processos=tipos_processos)

@bp_app.route("/create_categoria", methods=["POST"])
def create_categoria():
    form = dict(request.form)
    form["tipo_processo_nome"] = TipoProcessoController.get(form["tipo_processo"]).nome
    try:
        if form["_id"]:
            CategoriaController.update(form)
        else:
            CategoriaController.create(form)
        return "ok"
    except Exception as e:
        return (f"erro:{e}")

@bp_app.route("/delete_categoria/<id>")
def delete_categoria(id):
    try:
        CategoriaController.delete(id)
        return "ok"
    except Exception as e:
        return (f"erro:{e}")

@bp_app.route("/list_categorias/")
def list_categorias():
    categorias = CategoriaController.get_all()
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("list_categorias.html", categorias=categorias, permissao=permissao)

@bp_app.route("/tipo_processo_menu")
def tipo_processo_menu():
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("menu_tipo_processo.html", permissao=permissao)

@bp_app.route("/create_tipo_processo", methods=["POST"])
def create_tipo_processo():
    form = request.form
    try:
        if form["_id"]:
            TipoProcessoController.update(form)
        else:
            TipoProcessoController.create(form)
        return "ok"
    except Exception as e:
        return f"erro:{e}"

@bp_app.route("/delete_tipo_processo/<id>")
def delete_tipo_processo(id):
    try:
        TipoProcessoController.delete(id)
        return "ok"
    except Exception as e:
        return (f"erro:{e}")

@bp_app.route("/list_tipo_processo/")
def list_tipo_processo():
    tipo_processos = TipoProcessoController.get_all()
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("list_tipo_processo.html", tipo_processos=tipo_processos, permissao=permissao)

@bp_app.route("/perfil_menu")
def perfil_menu():
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("menu_perfil.html", permissao=permissao)

@bp_app.route("/create_perfil", methods=["POST"])
def create_perfil():
    form = request.form
    try:
        if form["_id"]:
            PerfilController.update(form)
        else:
            PerfilController.create(form)
        return "ok"
    except Exception as e:
        return f"erro:{e}"

@bp_app.route("/delete_perfil/<id>")
def delete_perfil(id):
    try:
        PerfilController.delete(id)
        return "ok"
    except Exception as e:
        return (f"erro:{e}")

@bp_app.route("/list_perfil/")
def list_perfil():
    perfis = PerfilController.get_all()
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("list_perfil.html", perfis=perfis, permissao=permissao)

@bp_app.route("/users_menu")
def users_menu():
    return render_template("menu_user.html")

@bp_app.route("/update_user/<user_id>/<perfil_id>")
def update_user(user_id, perfil_id):
    try:
        if user_id and perfil_id:
            UserController.update(user_id, perfil_id)
            return "ok"
    except Exception as e:
        return (f"erro:{e}")

@bp_app.route("/delete_user/<id>")
def delete_user(id):
    try:
        UserController.delete(id)
        return "ok"
    except Exception as e:
        return (f"erro:{e}")

@bp_app.route("/list_users/")
def list_users():
    users = UserController.get_all()
    perfis = PerfilController.get_all()
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("list_user.html", users=users, perfis=perfis,
                           permissao=permissao)

@bp_app.route("/adm_logs")
def adm_logs():
    return render_template("adm_logs.html")

@bp_app.route("/list_logs/")
def list_logs():
    logs = LogController.get_all()
    return render_template("list_logs.html", logs=logs)

@bp_app.route("/adm_cubo")
def adm_cubo():
    perfil_id = UserController.get(session["id"]).perfil_id
    permissao = PerfilController.get(perfil_id)
    return render_template("adm_cubo.html", permissao=permissao)

@bp_app.route("/adm_empresas")
def adm_empresas():
    return render_template("adm_empresas.html")

@bp_app.route("/adm_documentos")
def adm_documentos():
    permissao = PerfilController.get(UserController.get(session["id"]).perfil_id)
    uploaded = DocsController.is_all_uploaded()
    return render_template("adm_documentos.html", permissao=permissao, uploaded=uploaded)

@bp_app.route("/update_status/<_id>/<string:status>/", defaults={"obs": ""})
@bp_app.route("/update_status/<_id>/<string:status>/<obs>")
def update_status(_id, status, obs):
        DocsController.update_status(_id, obs, status)
        
        print(f"Updating document {_id} to status {status} with obs: {obs}")
        aprovacao = Aprovacao.query.get(_id)
        
        if aprovacao:
            doc = DocsController.get(aprovacao.documento_id)
        
        if doc.status == "APROVADO" or doc.status == "NAO APROVADO":
            return "ok|all"
        else:
            return "ok"

@bp_app.route("/upload_gdrive/")
def upload_gdrive():
        DocsController.upload_drive()
        
        return "ok"

@bp_app.route("/update_integra")
def update_integra():
    try:
        IntegraController.update_integracoes()
        return "ok"
    except Exception as e:
        return(f"erro:{e}")

@bp_app.route("/agendamento_integracao")
def agendamento_integracao():
    return render_template("menu_agendamento_integracao.html")

@bp_app.route("/list_agendamento_integracao/", defaults={"filter": ""})
@bp_app.route("/list_agendamento_integracao/<filter>")
def list_agendamento_integracao(filter):
    agendamentos = IntegraController.get_all_agendamento(filter)
    return render_template("list_agendamento_integracao.html", agendamentos=agendamentos)

@bp_app.route("/create_agendamento_integracao", methods=["POST"])
def create_agendamento_integracao():
    try:
        form = request.form
        if form["_id"]:
            IntegraController.update_agendamento(form)
        else:
            IntegraController.create_agendamento(form)
        return "ok"
    except Exception as e:
        return f"erro:{e}"

@bp_app.route("/cadastro_subcontratados")
def cadastro_subcontratados():
    return render_template("menu_subcontratados.html")

@bp_app.route("/list_subcontratados/")
def list_subcontratados():
    subcontratados = IntegraController.get_all_subcontratados()
    return render_template("list_subcontratados.html", subcontratados=subcontratados)

@bp_app.route("/create_subcontratados", methods=["POST"])
def create_subcontratados():
    try:
        form = request.form
        if form["_id"]:
            IntegraController.update_subcontratados(form)
        else:
            IntegraController.create_subcontratados(form)
        return "ok"
    
    except Exception as e:
        return f"erro:{e}"

@bp_app.route("/delete_subcontratados/<_id>")
def delete_subcontratados(_id):
    try:
        IntegraController.delete_subcontratados(_id)
        return "ok"
    except Exception as e:
        return f"erro:{e}"

@bp_app.route("/cadastro_funcionarios")
def cadastro_funcionarios():
    empresa = Empresa.query.filter_by(chave=session["chave_empresa"]).first()
    empresas_sub = Empresa.query.filter_by(_id=empresa._id).all()
    return render_template("menu_funcionarios.html", empresas=empresas_sub, empresa=empresa)

@bp_app.route("/adm_funcionarios")
def adm_funcionarios():
    empresas = EmpresaController.get_all()
    return render_template("adm_funcionarios.html", empresas=empresas)

@bp_app.route("/create_funcionario", methods=["POST"])
def create_funcionario():
    form = request.form
    if form["_id"]:
        if form["tipo"] == "status":
            rt = IntegraController.update_funcionario(form)
        else:
            rt = IntegraController.create_integra(form)
    else:
        rt = IntegraController.create_funcionario(form)
    return rt if rt else "ok"

@bp_app.route('/delete_funcionarios', methods=['POST'])
def delete_funcionarios():
    try:
        ids = request.get_json().get('ids', [])
        if not ids:
            return 'Nenhum ID recebido', 400
        from app.controllers.IntegraController import IntegraController
        deleted = 0
        for _id in ids:
            try:
                IntegraController.delete_funcionario(_id)
                deleted += 1
            except Exception as e:
                print(f'Erro ao excluir funcionário {_id}: {e}')
        if deleted:
            return 'ok'
        else:
            return 'Nenhum funcionário excluído', 400
    except Exception as e:
        print(f'Erro no endpoint delete_funcionarios: {e}')
        return 'Erro ao excluir funcionários', 500

@bp_app.route("/list_funcionarios/")
def list_funcionarios():
    if "id" in session.keys():
        funcionarios = IntegraController.get_all_funcionario()
        permissao = PerfilController.get(UserController.get(session["id"]).perfil_id)
    else:
        empresa_id = EmpresaController.get(chave=session["chave_empresa"])._id
        funcionarios = IntegraController.get_all_funcionario(empresa_id)
        permissao = False
    return render_template("list_funcionarios.html", funcionarios=funcionarios, permissao=permissao)

@bp_app.route("/menu_funcionario/<id>")
def menu_funcionario(id):        
        funcionario = IntegraController.get_funcionario(id)
        if not funcionario:
            return "Funcionário não encontrado", 404
        
        funcionarios = IntegraController.get_all_funcionario()
        
        empresa_id = StatusFuncionario.query.filter_by(funcionario_id=funcionario._id).order_by(StatusFuncionario.versao.desc()).first().empresa_id
        contratos = Contrato.query.filter_by(empresa_id=empresa_id).all()
        LogController.create(
            session.get("nome", "N/A"),
            session.get("perfil", "N/A"),
            "FUNCIONARIOS",
            "VIEW",
            f"Viewed menu for funcionário: {funcionario.nome}"
        )
        
        print(funcionario._id)
        
        return render_template(
            "modal_funcionario.html",
            contratos=contratos,
            funcionario=funcionario,
            funcionarios=funcionarios
        )

@bp_app.route("/aprovar_integracao/<_id>")
def aprovar_integracao(_id):
    try:
        if "id" not in session.keys():
            return "Unauthorized", 401
        if "SEGURANCA" not in session.get("perfil", ""):
            return "Insufficient permissions", 403
        funcionario = IntegraController.get_funcionario(_id)
        if not funcionario:
            return "Funcionário não encontrado", 404
        
        status = StatusFuncionario.query.filter_by(funcionario_id=funcionario._id, tipo="INTEGRACAO").order_by(StatusFuncionario.versao.desc()).first()
        
        if status:
            response = IntegraController.aprova(funcionario)
        else:
            return "Funcionário não possui integração pendente", 400
        
        if response.startswith("ok"):
            LogController.create(
                session.get("nome", "N/A"),
                session.get("perfil", "N/A"),
                "FUNCIONARIOS",
                "APPROVE",
                f"Approved integration for funcionário: {funcionario.nome}"
            )
            
            return response
        else:
            return response, 400
    
    except Exception as e:
        return f"Error approving integration: {str(e)}", 500

@bp_app.route("/download/<anexo_id>")
def download(anexo_id):
    anexo = Anexo.query.filter_by(_id=anexo_id).first()
    if not "perfil" in session.keys():
        abort(404, description="Não autorizado")
    
    if anexo is None:
        abort(404, description="Documento não encontrado")
    
    temp_file = BytesIO(zlib.decompress(anexo.data))
    return send_file(temp_file, download_name=f"{anexo.filename}.pdf")

@bp_app.route("/adm_relatorios")
def adm_relatorios():
    return render_template("adm_relatorios.html")


# PDF export route for integrations (regular users only)
@bp_app.route("/export_integracoes_pdf", methods=["GET"])
def export_integracoes_pdf():
    # Permission: only regular users (not PRESTADORA)
    if "perfil" not in session or session["perfil"] == "PRESTADORA":
        return "Acesso negado", 403

    # Get filters from query params
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")
    empresa_id = request.args.get("empresa_id")
    unidade_integracao = request.args.get("unidade_integracao")

    # Validate dates
    try:
        dt_inicio = datetime.strptime(data_inicio, "%Y-%m-%d") if data_inicio else None
        dt_fim = datetime.strptime(data_fim, "%Y-%m-%d") if data_fim else None
    except Exception:
        return "Formato de data inválido", 400

    # Query integrations
    integracoes = IntegraController.get_integracoes_pdf(dt_inicio, dt_fim, empresa_id, unidade_integracao)
    empresa_nome = EmpresaController.get(_id=empresa_id).nome if empresa_id else "Todas"

    # Gerar PDF horizontal (landscape)
    from reportlab.lib.pagesizes import landscape
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(30, height - 40, f"Relatório de Integrações")
    pdf.setFont("Helvetica", 12)
    pdf.drawString(30, height - 65, f"Empresa: {empresa_nome}")
    if dt_inicio and dt_fim:
        pdf.drawString(30, height - 85, f"Período: {dt_inicio.strftime('%d/%m/%Y')} a {dt_fim.strftime('%d/%m/%Y')}")
    
    pdf.line(30, height - 95, width - 30, height - 95)

    # Cabeçalho da tabela
    y = height - 120
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(30, y, "Nome")
    pdf.drawString(180, y, "Empresa")
    pdf.drawString(330, y, "Função")
    pdf.drawString(480, y, "Setor")
    pdf.drawString(630, y, "Data Integração")

    y -= 18
    pdf.setFont("Helvetica", 10)

    # Linhas da tabela
    for integra in integracoes:
        empresa_nome = integra.get("empresa_nome", "")
        # Se empresa_nome estiver vazio, tenta buscar pelo empresa_id
        print(empresa_nome)
        
        pdf.drawString(30, y, str(integra.get("nome", "")))
        pdf.drawString(180, y, str(empresa_nome))
        pdf.drawString(330, y, str(integra.get("funcao", "")))
        pdf.drawString(480, y, str(integra.get("setor", "")))
        # Data de integração isolada, espaçamento maior
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(630, y, str(integra.get("data_integracao", "")))
        pdf.setFont("Helvetica", 10)
        y -= 22
        if y < 100:
            pdf.showPage()
            y = height - 40

    # Campos de assinatura
    y -= 40
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(30, y, "Assinaturas dos Departamentos:")
    y -= 30
    pdf.setFont("Helvetica", 11)
    for dept in ["TI", "RH", "SEGURANÇA", "MEIO AMBIENTE", "RSC"]:
        pdf.drawString(30, y, f"{dept}: _______________________________")
        y -= 25

    pdf.save()
    buffer.seek(0)
    name = f"relatorio_integracoes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buffer, download_name=name, as_attachment=True)

# Excel export route for employee list (regular users only)
@bp_app.route("/export_funcionarios_excel", methods=["GET"])
def export_funcionarios_excel():
    if "perfil" not in session or session["perfil"] == "PRESTADORA":
        return "Acesso negado", 403

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")
    empresa_id = request.args.get("empresa_id")
    unidade_integracao = request.args.get("unidade_integracao")

    try:
        dt_inicio = datetime.strptime(data_inicio, "%Y-%m-%d") if data_inicio else None
        dt_fim = datetime.strptime(data_fim, "%Y-%m-%d") if data_fim else None
    except Exception:
        return "Formato de data inválido", 400

    # Query employees by filters
    funcionarios = IntegraController.get_funcionarios_excel(dt_inicio, dt_fim, empresa_id, unidade_integracao)
    if not funcionarios:
        return "Nenhum funcionário encontrado para os filtros selecionados.", 404
    
    empresa_nome = EmpresaController.get(_id=empresa_id).nome if empresa_id else "Todas"

    wb = Workbook()
    ws = wb.active
    ws.title = "Funcionários"

    # Header
    headers = ["Nome", "Função", "Cargo", "Setor", "Status", "Status Integração", "Empresa", "Data Integração", "Contrato", "CNPJ"]
    ws.append(headers)

    # Rows
    for f in funcionarios:
        ws.append([
            f.get("nome", ""),
            f.get("funcao", ""),
            f.get("cargo", ""),
            f.get("setor", ""),
            f.get("status", ""),
            f.get("status_integracao", ""),
            f.get("empresa_nome", ""),
            f.get("data_integracao", ""),
            f.get("contrato", ""),
            f.get("empresa_cnpj", "")
        ])

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"funcionarios_{empresa_nome}.xlsx"
    return send_file(buffer, download_name=filename, as_attachment=True)

# Endpoint para upload dos documentos obrigatórios do funcionário
@bp_app.route("/upload_documentos_funcionario", methods=["POST"])
def upload_documentos_funcionario():
    files = request.files
    funcionario_id = request.form.get("status_funcionario_id")

    # Se quiser associar ao último status, pode buscar pelo funcionário
    # status_funcionario_id = ...
    required_docs = [
        ("ficha_registro", "Cópia de ficha de registro de empregado"),
        ("registro_esocial", "Registro de e-social"),
        ("aso", "Cópia do ASO"),
        ("os", "O.S"),
        ("apr", "A.P.R"),
        ("ficha_epi", "Ficha de EPI")
    ]
    
    anexos = []
    for field, label in required_docs:
        file = files.get(field)
        if not file:
            return f"Documento obrigatório não enviado: {label}", 400
        
        data = file.read()
        file_extension = os.path.splitext(secure_filename(file.filename))[1].lower()
        
        anexo = AnexosFuncionario(
            filename=f"{field}.{file_extension.lstrip('.')}",
            funcionario_id= funcionario_id if funcionario_id else 0,
            data=data,
            link="",
            corrigido=False,
            hash="" # Pode gerar hash do arquivo se necessário
        )
        
        anexos.append(anexo)
    try:
        for anexo in anexos:
            db.session.add(anexo)
            db.session.flush()  # Gera o ID do anexo
            
            anexo.link = url_for("bp.download_doc_funcionario", anexo_id=anexo._id)
            db.session.commit()
                    
        return "ok"
    except Exception as e:
        db.session.rollback()
        return f"Erro ao salvar documentos: {str(e)}", 500

# Endpoint para retornar lista de documentos obrigatórios do funcionário e status/link
@bp_app.route("/api/documentos_funcionario/<funcionario_id>", methods=["GET"])
def api_documentos_funcionario(funcionario_id):
    # Lista de documentos obrigatórios
    required_docs = [
        ("ficha_registro", "Cópia de ficha de registro de empregado"),
        ("registro_esocial", "Registro de e-social"),
        ("aso", "Cópia do ASO"),
        ("os", "O.S"),
        ("apr", "A.P.R"),
        ("ficha_epi", "Ficha de EPI")
    ]
    
    # Buscar anexos já enviados para o funcionário
    anexos = AnexosFuncionario.query.filter_by(funcionario_id=funcionario_id).all()
    documentos = []
    for chave, nome in required_docs:
        doc = {
            "chave": chave,
            "nome": nome,
            "url": None
        }
        
        # Verifica se já existe o documento enviado
        for anexo in anexos:
            if chave in anexo.filename:
                doc["url"] = anexo.link
                break
        documentos.append(doc)
    
    return jsonify({"documentos": documentos})

@bp_app.route("/download_doc_funcionario/<anexo_id>")
def download_doc_funcionario(anexo_id):
    anexo = AnexosFuncionario.query.filter_by(_id=anexo_id).first()
    if anexo is None:
        abort(404, description="Documento não encontrado")
    
    temp_file = BytesIO(anexo.data)
    return send_file(temp_file, download_name=anexo.filename, as_attachment=True)

# Register blueprint
def configure(app):
    app.register_blueprint(bp_app)
